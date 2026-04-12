"""Parse admin Excel uploads for shared lead pool (column order / headers flexible)."""

from __future__ import annotations

import re
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

# Expected column semantics (first row headers). Order in sheet can differ.
_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "submit_time": ("submit time", "submitted at", "date", "timestamp"),
    "name": ("full name", "name", "lead name"),
    "age": ("age",),
    "gender": ("gender", "sex"),
    "phone": (
        "phone number (calling number)",
        "phone number",
        "calling number",
        "mobile",
        "phone",
    ),
    "city": ("your city name", "city", "city name"),
    "ad_name": ("ad name", "ad", "ad id"),
}


def _norm_header(cell: Any) -> str:
    if cell is None:
        return ""
    s = str(cell).strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _map_headers(row: tuple[Any, ...]) -> dict[str, int]:
    """Map semantic key -> 0-based column index (best header match per key)."""
    headers = [_norm_header(c) for c in row]
    best: dict[str, tuple[int, int]] = {}
    for col_i, h in enumerate(headers):
        if not h:
            continue
        for key, aliases in _HEADER_ALIASES.items():
            score = 0
            for a in aliases:
                if h == a:
                    score = max(score, 200 + len(a))
                elif h.startswith(a) or h.startswith(a + " ") or h.startswith(a + "("):
                    score = max(score, 150 + len(a))
                elif len(a) >= 6 and a in h:
                    score = max(score, 80 + len(a))
                elif len(a) >= 4 and a in h:
                    score = max(score, 40 + len(a))
            if score > 0:
                prev = best.get(key)
                if prev is None or score > prev[0]:
                    best[key] = (score, col_i)
    return {k: v[1] for k, v in best.items()}


def _cell_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _parse_age(val: Any) -> int | None:
    if val is None or val == "":
        return None
    try:
        if isinstance(val, (int, float)):
            a = int(val)
        else:
            s = re.sub(r"\D", "", str(val))
            if not s:
                return None
            a = int(s)
        if 0 < a < 130:
            return a
    except (ValueError, TypeError):
        return None
    return None


def _parse_submit_time(val: Any) -> datetime | None:
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        dt = val
        if dt.tzinfo is None:
            return dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(timezone.utc)
    if isinstance(val, date) and not isinstance(val, datetime):
        return datetime(val.year, val.month, val.day, tzinfo=timezone.utc)
    if isinstance(val, (int, float)):
        try:
            dt = from_excel(val)
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            else:
                dt = dt.astimezone(timezone.utc)
            return dt
        except Exception:
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            if dt.tzinfo is None:
                dt = dt.replace(tzinfo=timezone.utc)
            return dt.astimezone(timezone.utc)
        except ValueError:
            return None
    return None


def _normalize_phone(raw: str) -> str:
    digits = re.sub(r"\D", "", raw)
    if len(digits) >= 10:
        return digits[-10:]
    return digits


def parse_pool_xlsx_rows(content: bytes) -> tuple[list[dict[str, Any]], list[str]]:
    """Return (row dicts, parse warnings). Each dict: name, phone, city, age, gender, ad_name, submit_time."""
    warnings: list[str] = []
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        header_row = next(it, None)
        if not header_row:
            return [], ["Empty spreadsheet"]
        colmap = _map_headers(tuple(header_row))
        if "name" not in colmap:
            return [], [
                "Could not find a “Full Name” column. Use headers like: Submit Time, Full Name, Age, …",
            ]

        out: list[dict[str, Any]] = []
        for row in it:
            if not row:
                continue
            def get(key: str) -> Any:
                j = colmap.get(key)
                if j is None or j >= len(row):
                    return None
                return row[j]

            name = _cell_str(get("name"))
            if not name:
                continue

            pn = get("phone")
            if isinstance(pn, (int, float)) and not isinstance(pn, bool):
                phone_raw = str(int(pn))
            else:
                phone_raw = _cell_str(pn)
            city = _cell_str(get("city")) or None
            gender = _cell_str(get("gender")) or None
            if gender:
                gender = gender[:32]
            ad_name = _cell_str(get("ad_name")) or None
            if ad_name:
                ad_name = ad_name[:255]

            st = _parse_submit_time(get("submit_time"))

            out.append(
                {
                    "name": name[:255],
                    "phone": _normalize_phone(phone_raw) or None,
                    "city": city,
                    "age": _parse_age(get("age")),
                    "gender": gender,
                    "ad_name": ad_name,
                    "submit_time": st,
                }
            )
        if not out:
            warnings.append("No data rows with a non-empty Full Name were found.")
        return out, warnings
    finally:
        wb.close()
