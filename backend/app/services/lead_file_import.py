"""Parse PDF or Excel (.xlsx) for team & leader personal lead imports (calling board).

PDF table / text extraction matches legacy ``myle_dashboard`` ``_extract_leads_from_pdf``.
Excel imports read a header row (Name / Phone / Email / City in any column order).
"""

from __future__ import annotations

import io
import re
from dataclasses import dataclass

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.activity_log import ActivityLog
from app.models.lead import Lead
from app.schemas.leads import LeadCreate
from app.services.auto_handoff import AutoHandoffService
from app.services.crm_outbox import enqueue_lead_shadow_upsert

_PHONE_RE = re.compile(r"(?:(?:\+|0{0,2})91[-\s]?)?([6-9]\d{9})\b")
_MAX_BYTES = 5 * 1024 * 1024
_SOURCE_OK = frozenset({"facebook", "instagram", "referral", "walk_in", "other"})


def normalize_phone_digits(raw: str) -> str | None:
    digits = re.sub(r"\D", "", raw or "")
    if len(digits) >= 10:
        return digits[-10:]
    return None


def extract_leads_from_pdf_bytes(content: bytes) -> tuple[list[dict[str, str | None]], str | None]:
    """Return (rows, error_message). Each row: name, phone, email, city, source=None, extra_notes=None."""
    try:
        import pdfplumber
    except ImportError:
        return [], "PDF parsing is not available on this server (pdfplumber not installed)."

    leads: list[dict[str, str | None]] = []
    try:
        with pdfplumber.open(io.BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                if tables:
                    for table in tables:
                        if not table:
                            continue
                        header_row = [str(c or "").lower().strip() for c in table[0]]
                        name_col = next((i for i, h in enumerate(header_row) if "name" in h), None)
                        phone_col = next(
                            (
                                i
                                for i, h in enumerate(header_row)
                                if any(k in h for k in ("phone", "mobile", "contact", "number"))
                            ),
                            None,
                        )
                        email_col = next(
                            (i for i, h in enumerate(header_row) if "email" in h or "mail" in h),
                            None,
                        )
                        city_col = next(
                            (i for i, h in enumerate(header_row) if "city" in h or "location" in h),
                            None,
                        )
                        start = 1 if (name_col is not None or phone_col is not None) else 0
                        for row in table[start:]:
                            if not row:
                                continue
                            cells = [str(c or "").strip() for c in row]

                            def safe(i: int | None) -> str:
                                if i is None or i >= len(cells):
                                    return ""
                                return cells[i]

                            name = safe(name_col)
                            phone = safe(phone_col)
                            email = safe(email_col) or None
                            city = safe(city_col) or None
                            m = _PHONE_RE.search(phone)
                            if m:
                                phone = m.group(1)
                            if name or phone:
                                leads.append(
                                    {
                                        "name": name,
                                        "phone": phone,
                                        "email": email,
                                        "city": city,
                                        "source": None,
                                        "extra_notes": None,
                                    }
                                )
                else:
                    text = page.extract_text() or ""
                    for line in text.split("\n"):
                        m = _PHONE_RE.search(line)
                        if not m:
                            continue
                        phone = m.group(1)
                        name = _PHONE_RE.sub("", line).strip(" -|,;:\t")
                        leads.append(
                            {
                                "name": name,
                                "phone": phone,
                                "email": None,
                                "city": None,
                                "source": None,
                                "extra_notes": None,
                            }
                        )
    except Exception as exc:
        return [], f"Could not parse PDF: {exc}"

    return leads, None


# Header keyword -> semantic field for spreadsheet (xlsx) imports. First keyword
# that appears in a header cell wins for that column.
_XLSX_FIELD_KEYWORDS: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("email", ("email", "mail")),
    ("phone", ("phone", "mobile", "contact", "number", "calling")),
    ("city", ("city", "location")),
    ("name", ("name",)),
)


def _cell_text(val: object) -> str:
    if val is None:
        return ""
    if isinstance(val, bool):
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def extract_leads_from_xlsx_bytes(content: bytes) -> tuple[list[dict[str, str | None]], str | None]:
    """Return (rows, error_message) from an .xlsx workbook.

    Rows match the PDF extractor shape: name, phone, email, city, source=None, extra_notes=None.
    Headers are matched by keyword so column order does not matter.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [], "Excel parsing is not available on this server (openpyxl not installed)."

    leads: list[dict[str, str | None]] = []
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        return [], f"Could not parse Excel file: {exc}"
    try:
        ws = wb.active
        if ws is None:
            return [], "The Excel file has no readable sheet."
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, None)
        if not header_row:
            return [], "The Excel file is empty."

        colmap: dict[str, int] = {}
        for col_i, cell in enumerate(header_row):
            h = _cell_text(cell).lower()
            if not h:
                continue
            for field, keywords in _XLSX_FIELD_KEYWORDS:
                if field in colmap:
                    continue
                if any(k in h for k in keywords):
                    colmap[field] = col_i
                    break

        if "name" not in colmap and "phone" not in colmap:
            return [], (
                "Could not find a Name or Phone column. Use a header row like: Name, Phone, Email, City."
            )

        for row in rows_iter:
            if not row:
                continue

            def cell(field: str) -> str:
                j = colmap.get(field)
                if j is None or j >= len(row):
                    return ""
                return _cell_text(row[j])

            name = cell("name")
            phone = cell("phone")
            email = cell("email") or None
            city = cell("city") or None
            m = _PHONE_RE.search(phone)
            if m:
                phone = m.group(1)
            if name or phone:
                leads.append(
                    {
                        "name": name,
                        "phone": phone,
                        "email": email,
                        "city": city,
                        "source": None,
                        "extra_notes": None,
                    }
                )
    finally:
        wb.close()

    return leads, None


def _resolve_source(row_src: str | None, fallback_tag: str) -> str:
    s = (row_src or "").strip().lower()
    if s in _SOURCE_OK:
        return s
    t = fallback_tag.strip().lower()
    if t in _SOURCE_OK:
        return t
    return "other"


def _build_notes(filename: str, tag: str, extra: str | None) -> str | None:
    parts = [f"Imported from {filename}"]
    if tag.strip():
        parts.append(f"Tag: {tag.strip()}")
    if extra:
        parts.append(extra)
    joined = " — ".join(parts)
    return joined[:5000] if joined else None


@dataclass
class LeadImportResult:
    imported: int
    skipped: int
    warnings: list[str]


async def run_personal_lead_import(
    session: AsyncSession,
    *,
    user_id: int,
    file_bytes: bytes,
    filename: str,
    source_tag: str,
) -> LeadImportResult:
    """Insert leads for ``user_id`` (creator + assignee). Skip duplicates by normalized phone."""
    if len(file_bytes) > _MAX_BYTES:
        return LeadImportResult(0, 0, [f"File too large (max {_MAX_BYTES // (1024 * 1024)} MB)."])

    fname = (filename or "").lower()
    warnings: list[str] = []

    if fname.endswith(".pdf"):
        rows, err = extract_leads_from_pdf_bytes(file_bytes)
    elif fname.endswith(".xlsx"):
        rows, err = extract_leads_from_xlsx_bytes(file_bytes)
    else:
        return LeadImportResult(0, 0, ["Only .pdf and .xlsx files are allowed."])

    if err:
        return LeadImportResult(0, 0, [err])

    phones_db = (
        await session.execute(select(Lead.phone).where(Lead.deleted_at.is_(None)))
    ).scalars().all()
    existing: set[str] = set()
    for p in phones_db:
        n = normalize_phone_digits(p or "")
        if n and len(n) == 10:
            existing.add(n)

    imported = 0
    skipped = 0
    seen_in_file: set[str] = set()

    for row in rows:
        name = (row.get("name") or "").strip()
        phone_raw = (row.get("phone") or "").strip()
        norm = normalize_phone_digits(phone_raw)
        if not name and not phone_raw:
            skipped += 1
            continue
        if not norm or len(norm) != 10:
            skipped += 1
            continue
        if norm in existing or norm in seen_in_file:
            skipped += 1
            continue
        seen_in_file.add(norm)
        if not name:
            name = norm
        src = _resolve_source(row.get("source"), source_tag)
        extra = row.get("extra_notes")
        if isinstance(extra, str):
            extra = extra.strip() or None
        notes = _build_notes(filename or "upload", source_tag, extra)
        body = LeadCreate(
            name=name[:255],
            status="new_lead",
            phone=norm,
            email=(row.get("email") or None),
            city=(row.get("city") or None),
            source=src,
            notes=notes,
        )
        lead = Lead(
            name=body.name,
            status=body.status,
            created_by_user_id=user_id,
            owner_user_id=user_id,
            assigned_to_user_id=user_id,
            phone=body.phone,
            email=body.email,
            city=body.city,
            source=body.source,
            notes=body.notes,
        )
        session.add(lead)
        await session.flush()
        handoff = AutoHandoffService(session)
        await handoff.on_lead_created(lead=lead, actor_user_id=user_id)
        session.add(
            ActivityLog(
                user_id=user_id,
                action="lead.created",
                entity_type="lead",
                entity_id=lead.id,
                meta={"name": lead.name, "status": lead.status, "via": "file_import"},
            ),
        )
        enqueue_lead_shadow_upsert(session, lead)
        existing.add(norm)
        imported += 1

    await session.commit()
    if imported == 0 and not warnings:
        warnings.append("No new leads were imported (empty file, bad phone numbers, or all duplicates).")
    return LeadImportResult(imported=imported, skipped=skipped, warnings=warnings)
